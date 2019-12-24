import openpyxl as xl
import random
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb['Sheet1']

    for i in range(2, 52):
        for j in range(3, 8):
            cell = sheet.cell(i, j)
            cell.value = random.randint(0, 100)

        total_value_cell = sheet.cell(i, 8)
        total_value_cell.value = sheet.cell(i, 3).value + sheet.cell(i, 4).value + sheet.cell(i, 5).value + sheet.cell(
            i, 6).value + sheet.cell(i, 7).value
        avg_value_cell = sheet.cell(i, 9)
        avg_value_cell.value = total_value_cell.value / 5

    values = Reference(sheet, min_row=2, max_row=51, min_col=9, max_col=9)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'j2')

    wb.save(filename)

filename='SEM_1.xlsx'
process_workbook(filename)

filename='SEM_2.xlsx'
process_workbook(filename)

filename='SEM_3.xlsx'
process_workbook(filename)

filename='SEM_4.xlsx'
process_workbook(filename)

filename='SEM_5.xlsx'
process_workbook(filename)

